from pathlib import Path

import napari
import numpy as np
import openpyxl
from magicgui import magic_factory
from napari import layers
from napari.utils.notifications import show_info, show_warning, show_error
from skimage import measure, filters, restoration

from scipy.ndimage import label

from .segmentation import watershed_pores, get_particle_border_zone


@magic_factory(
    call_button='Check',
    labels={'label': 'Labels'},
)
def check_disconnected_regions(
        labels: layers.Labels
):
    if labels is None:
        show_warning('No layer selected')
        return

    last_label = np.max(np.unique(labels.data))
    empty_labels: list[int] = []
    disconnected_labels: list[int] = []

    for i in range(1, last_label + 1):
        binary_mask = (labels.data == i)
        labeled_mask = measure.label(binary_mask)
        regions = measure.regionprops(labeled_mask)

        if len(regions) == 0:
            empty_labels.append(i)
        elif len(regions) > 1:
            disconnected_labels.append(i)

    show_info(f'Last non-empty label: {last_label}')
    show_info(f'{len(empty_labels)} empty labels: {", ".join(map(str, empty_labels))}')
    show_info(f'{len(disconnected_labels)} disconnected labels: {", ".join(map(str, disconnected_labels))}')


@magic_factory(
    call_button="Clean",
    labels={"label": "Labels"},
)
def clean_disconnected_regions(labels: layers.Labels):
    data = np.asarray(labels.data)
    cleaned = data.copy()

    unique_ids = np.unique(data)
    unique_ids = unique_ids[unique_ids != 0]

    n_labels_cleaned: int = 0
    n_regions_cleaned: int = 0
    for lbl in unique_ids:
        mask = data == lbl

        labeled_mask, n_comps = measure.label(
            mask, return_num=True, connectivity=1
        )

        if n_comps <= 1:
            continue

        props = measure.regionprops(labeled_mask)
        n_regions_before = n_regions_cleaned
        for region in props:
            if region.area < 20:
                # coords is (N, ndim) array – can be used directly for indexing
                cleaned[tuple(region.coords.T)] = 0
                n_regions_cleaned += 1

        if n_regions_cleaned > n_regions_before:
            n_labels_cleaned += 1

    # Write the cleaned data back to the layer (updates the viewer automatically)
    labels.data = cleaned
    show_info(f'Cleaned a total of {n_labels_cleaned} labels, removed {n_regions_cleaned} tiny regions')


@magic_factory(
    call_button='Get pore area fraction',
    labels={'label': 'Particle labels'},
    crop_border_distance={'label': 'Crop border, px', 'widget_type': 'Slider', 'min': 0, 'max': 5, 'step': 1},
)
def measure_pore_area_fraction(
        viewer: napari.Viewer,
        labels: layers.Labels,
        crop_border_distance: int = 3
):
    if labels is None:
        show_warning('No layer selected')
        return

    pore_data = viewer.layers['porous_structure'].data
    particle_mask = labels.data

    border_mask = get_particle_border_zone(particle_mask, crop_border_distance,
                                           background_border_only=False)
    particle_mask = (labels.data != 0)
    pore_mask = (pore_data == 1)
    # subtract border zone for less error
    pore_mask = pore_mask & (~border_mask)
    particle_mask = particle_mask & (~border_mask)

    fraction = np.sum(pore_mask) / np.sum(particle_mask)
    fraction_percent = round(fraction * 100, 2)

    show_info(f'Total pore area fraction: {fraction_percent}%')


@magic_factory(
    call_button='Export to spreadsheet',
    labels={'label': 'Particle labels'},
    conversion_factor={'widget_type': 'FloatSpinBox', 'min': 1, 'max': 1000, 'label': 'px/µm'},
    export_pore_area={'widget_type': 'CheckBox', 'value': True, 'text': 'Include pore area fraction'},
    output_format={'label': 'Export format: ', 'choices': ['.xlsx', '.csv']},
    output_path={
        'label': 'Save as...',
        'mode': 'w',  # 'w' for write mode (save new file)
        'filter': '*.csv *.xlsx'
    }
)
def export_data(
        viewer: napari.Viewer,
        labels: layers.Labels,
        conversion_factor: float,
        export_pore_area: bool,
        output_format,
        output_path: Path
):
    if (output_path is None) or (output_path == Path('')):
        show_warning('Choose an output path')
        return
    if labels is None:
        show_warning('No layer selected')
        return
    if export_pore_area and ('porous_structure' not in viewer.layers):
        show_warning('No "porous_structure" layer present.\nRun porous structure segmentation first to export this data.')
        return

    indexes: list[int] = []
    eq_diameters: list[float] = []
    feret_diameters: list[float] = []
    areas: list[float] = []
    pore_area_fractions: list[float] = []

    labels_data = labels.data
    last_label = np.max(np.unique(labels_data))

    if export_pore_area:
        pore_data = viewer.layers['porous_structure'].data
        pore_mask = (pore_data == 1)
        pore_area_counts = np.bincount(labels_data[pore_mask], minlength=last_label + 1)

    for i in range(1, last_label + 1):
        binary_mask = (labels_data == i)
        labeled_mask = measure.label(binary_mask)
        regions = measure.regionprops(labeled_mask)

        if len(regions) == 0:
            show_warning(f'Label {i} is empty, skipping')
            continue
        elif len(regions) > 1:
            show_warning(f'Label {i} is split into {len(regions)} disconnected regions.\n'
                         f'Cannot determine which one to measure.')
            continue

        eq_diameter_px: float = regions[0].equivalent_diameter_area
        eq_diameter: float = round(eq_diameter_px / conversion_factor, 2)

        feret_diameter_px: float = regions[0].feret_diameter_max
        feret_diameter: float = round(feret_diameter_px / conversion_factor, 2)

        area_px2: float = regions[0].area
        area: float = round(area_px2 / (conversion_factor ** 2), 2)

        eq_diameters.append(eq_diameter)
        feret_diameters.append(feret_diameter)
        areas.append(area)
        indexes.append(i)

        if export_pore_area:
            pore_area_px2 = pore_area_counts[i]
            pore_area_fractions.append(float(pore_area_px2) / float(area_px2))

    data_to_save = [indexes, eq_diameters, feret_diameters, areas]
    if export_pore_area:
        data_to_save.append(pore_area_fractions)
    data_to_save = np.column_stack(data_to_save)

    output_path = output_path.with_suffix(output_format)

    try:
        if output_format == '.csv':
            header = 'Label,Equivalent diameter (µm),Max Feret diameter (µm),Area (µm^2)'
            if export_pore_area:
                header += ',Pore area fraction'
            np.savetxt(str(output_path), data_to_save, delimiter=',', header=header)

        elif output_format == '.xlsx':
            wb = openpyxl.Workbook()
            sheet = wb.active

            sheet.title = 'Exported Data'
            sheet['A1'] = 'Label'
            sheet['B1'] = 'Equivalent diameter (µm)'
            sheet['C1'] = 'Max Feret diameter (µm)'
            sheet['D1'] = 'Area (µm^2)'
            if export_pore_area:
                sheet['E1'] = 'Pore area fraction'

            for row, values in enumerate(data_to_save, start=2):  # start=2 to skip header
                sheet.cell(row=row, column=1, value=values[0])
                sheet.cell(row=row, column=2, value=values[1])
                sheet.cell(row=row, column=3, value=values[2])
                sheet.cell(row=row, column=4, value=values[3])
                if export_pore_area:
                    sheet.cell(row=row, column=5, value=values[4])

            wb.save(str(output_path))

    except Exception as e:
        show_error(f'An error occured during export: {e}')
        return

    show_info('Successfully exported to: ' + str(output_path))


@magic_factory(
    call_button='Segment',
    image={'label': 'Image'},
    particle_labels={'label': 'Particle labels'},
    edge_filter={'label': 'Edge detection', 'choices': ['Sobel', 'Prewitt', 'Farid', 'Scharr']},
    min_area={'widget_type': 'Slider', 'min': 0, 'max': 50, 'step': 1, 'label': 'min area (px)'},
    max_area={'widget_type': 'Slider', 'min': 100, 'max': 500, 'step': 1, 'label': 'max area (px)'},
    min_depth={'widget_type': 'Slider', 'min': 10, 'max': 30, 'step': 1},
    min_border_distance={'widget_type': 'Slider', 'min': 0, 'max': 5, 'step': 1},
    background_border_only={'label': 'Only check background border'}
)
def segment_porous_structure(
        image: layers.Image,
        particle_labels: layers.Labels,
        edge_filter: str = 'Scharr',
        min_depth: int = 20,
        min_area: int = 10,
        max_area: int = 300,
        min_border_distance: int = 3,
        background_border_only: bool = True
) -> napari.types.LayerDataTuple:
    if image is None:
        show_warning('No image selected')
        return
    if particle_labels is None:
        show_warning('No labels layer selected')
        return

    img_data = image.data
    output_data = np.zeros_like(img_data, dtype=np.uint8)

    pore_mask, markers = watershed_pores(img_data,
                                         edge_filter=edge_filter,
                                         min_depth=min_depth,
                                         min_area=min_area,
                                         max_area=max_area)

    particle_mask = particle_labels.data
    pore_mask = pore_mask & (particle_mask != 0)  # pores only inside particles
    output_data[pore_mask] = 1

    border_zone = get_particle_border_zone(particle_mask, min_border_distance,
                                           background_border_only=background_border_only)

    # Remove pores intersecting the border zone
    pore_labels, num_pores = label(output_data)  # label current pores
    for label_id in range(1, num_pores + 1):
        pore_component = (pore_labels == label_id)
        if np.any(pore_component & border_zone):
            output_data[pore_component] = 0

    return (output_data, {'name': 'porous_structure'}, 'labels')


@magic_factory(
    call_button='Test segmentation',
    image={'label': 'Image'},
    particle_labels={'label': 'Particle labels'},
    marker_labels={'label': 'Layer for markers'},
    edge_filter={'label': 'Edge detection', 'choices': ['Sobel', 'Prewitt', 'Farid', 'Scharr']},
    min_area={'widget_type': 'Slider', 'min': 0, 'max': 50, 'step': 1, 'label': 'min area (px)'},
    max_area={'widget_type': 'Slider', 'min': 100, 'max': 500, 'step': 1, 'label': 'max area (px)'},
    min_depth={'widget_type': 'Slider', 'min': 10, 'max': 30, 'step': 1},
    min_border_distance={'widget_type': 'Slider', 'min': 0, 'max': 5, 'step': 1},
    background_border_only={'label': 'Only check background border'}
)
def test_segment_porous_structure(
        image: layers.Image,
        particle_labels: layers.Labels,
        marker_labels: layers.Labels,
        edge_filter: str = 'Scharr',
        min_depth: int = 20,
        min_area: int = 10,
        max_area: int = 300,
        min_border_distance: int = 3,
        background_border_only: bool = True
) -> napari.types.LayerDataTuple:
    img_data = image.data
    output_data = np.zeros_like(img_data, dtype=np.uint8)

    pore_mask, markers = watershed_pores(img_data,
                                         edge_filter=edge_filter,
                                         min_depth=min_depth,
                                         min_area=min_area,
                                         max_area=max_area)

    particle_mask = particle_labels.data
    pore_mask = pore_mask & (particle_mask != 0)  # pores only inside particles
    output_data[pore_mask] = 1

    border_zone = get_particle_border_zone(particle_mask, min_border_distance, background_border_only=background_border_only)

    # Remove pores intersecting the border zone
    pore_labels, num_pores = label(output_data)  # label current pores
    for label_id in range(1, num_pores + 1):
        pore_component = (pore_labels == label_id)
        if np.any(pore_component & border_zone):
            output_data[pore_component] = 0

    marker_labels.data = border_zone

    return (output_data, {'name': 'porous_structure'}, 'labels')


@magic_factory(
    call_button='Add figure to layers',
)
def export_figure(
        viewer: napari.Viewer
):
    figure = viewer.export_figure()
    viewer.add_image(figure, rgb=True, name='figure')
